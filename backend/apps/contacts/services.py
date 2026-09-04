import io
import re
from decimal import Decimal
from typing import List, Dict, Any
import openpyxl
import openpyxl.styles
from django.db import transaction, models
from django.utils import timezone
from django.core.exceptions import ValidationError
from apps.accounting.models import Account, AccountType, JournalEntry, ReferenceType
from apps.accounting.services import AccountingService
from apps.sales.models import Sale, SaleStatus, SalesReturn, PaymentMethodType
from .models import Customer, Supplier, CustomerPayment, CustomerPaymentStatus


class CustomerReceivableService:
    """
    Authoritative calculation engine for customer credit receivables, statements, and payments.
    """

    @classmethod
    def reallocate_customer_payments(cls, customer: Customer):
        """
        Allocates all submitted CustomerPayment records across the customer's completed credit sales in FIFO order.
        Updates each Sale's paid_amount and due_amount so that the Sales Invoices list stays 100% in sync with Customer Payments.
        """
        if customer.is_walkin:
            return

        sales = list(Sale.objects.filter(customer=customer, status=SaleStatus.COMPLETED).order_by("date", "id"))
        payments = list(CustomerPayment.objects.filter(customer=customer, status=CustomerPaymentStatus.SUBMITTED).order_by("date", "id"))

        total_payment_pool = sum(p.amount for p in payments)

        # 1. Reset each sale's initial paid and due amounts based on upfront payment (at checkout)
        for s in sales:
            upfront_paid = Decimal("0.00")
            if s.payment_method in [PaymentMethodType.CASH, PaymentMethodType.CARD, PaymentMethodType.CHEQUE]:
                upfront_paid = s.grand_total
            elif s.payments.exists():
                upfront_paid = sum(
                    p.amount for p in s.payments.filter(
                        payment_method__in=[PaymentMethodType.CASH, PaymentMethodType.CARD, PaymentMethodType.CHEQUE]
                    )
                )
            else:
                if s.payment_method != PaymentMethodType.CREDIT:
                    upfront_paid = min(s.paid_amount, s.grand_total)

            # Only deduct returns from invoice total if the return was an AR credit reduction (Credit Sale)
            if s.payment_method == PaymentMethodType.CREDIT:
                returns_amt = sum(r.refund_amount for r in s.returns.all())
                effective_grand_total = max(Decimal("0.00"), s.grand_total - returns_amt)
            else:
                effective_grand_total = s.grand_total

            s.paid_amount = min(upfront_paid, effective_grand_total)
            s.due_amount = max(Decimal("0.00"), effective_grand_total - s.paid_amount)

        # 2. Opening balance absorption
        base_opening = customer.opening_balance or Decimal("0.00")
        payment_pool_for_sales = max(Decimal("0.00"), total_payment_pool - base_opening)

        # 3. Allocate remaining payment pool across sales in FIFO order
        for s in sales:
            if s.due_amount > Decimal("0.00") and payment_pool_for_sales > Decimal("0.00"):
                alloc = min(payment_pool_for_sales, s.due_amount)
                s.paid_amount += alloc
                s.due_amount -= alloc
                payment_pool_for_sales -= alloc

            s.save(update_fields=["paid_amount", "due_amount", "updated_at"])

    @classmethod
    def get_customer_outstanding(cls, customer_id: int) -> dict:
        """
        Calculates authoritative outstanding receivable balance for a customer.
        Uses exact statement closing balance to ensure 100% synchronization across all views.
        """
        customer = Customer.objects.get(pk=customer_id)
        if customer.is_walkin:
            return {
                "customer_id": customer.id,
                "customer_code": customer.customer_id,
                "customer_name": customer.name,
                "is_walkin": True,
                "credit_enabled": False,
                "opening_balance": Decimal("0.00"),
                "total_credit_sales": Decimal("0.00"),
                "total_payments": Decimal("0.00"),
                "total_returns": Decimal("0.00"),
                "outstanding_balance": Decimal("0.00"),
            }

        statement_data = cls.get_customer_statement(customer_id)
        summary = statement_data["summary"]

        return {
            "customer_id": customer.id,
            "customer_code": customer.customer_id,
            "customer_name": customer.name,
            "is_walkin": False,
            "credit_enabled": customer.credit_enabled,
            "opening_balance": Decimal(str(summary["opening_balance"])),
            "total_credit_sales": Decimal(str(summary["total_sales"])),
            "total_payments": Decimal(str(summary["total_payments"])),
            "total_returns": Decimal(str(summary["total_returns"])),
            "outstanding_balance": Decimal(str(summary["closing_balance"])),
        }

    @classmethod
    @transaction.atomic
    def create_payment(cls, data: dict, user, submit_now: bool = True) -> CustomerPayment:
        """
        Creates and executes a Customer Payment transaction.
        Enforces overpayment prevention and posts double-entry accounting entries.
        """
        customer = data.get("customer")
        if isinstance(customer, int):
            customer = Customer.objects.get(pk=customer)

        if customer.is_walkin:
            raise ValidationError("Cannot record payments for the default Walk-in Customer.")

        amount = Decimal(str(data.get("amount", "0")))
        date = data.get("date") or timezone.localdate()
        if isinstance(date, str):
            from datetime import datetime
            try:
                date = datetime.strptime(date, "%Y-%m-%d").date()
            except Exception:
                date = timezone.localdate()

        # Prevent backdating payment before customer's latest unpaid invoice or current server date
        if customer.sales.filter(due_amount__gt=0).exists():
            latest_unpaid_date = customer.sales.filter(due_amount__gt=0).aggregate(m=models.Max("date"))["m"]
            if latest_unpaid_date and date < latest_unpaid_date:
                date = latest_unpaid_date
        if date < timezone.localdate():
            date = timezone.localdate()

        payment_method = data.get("payment_method", "CASH")
        payment_account = data.get("payment_account")
        reference = data.get("reference", "").strip()
        notes = data.get("notes", "").strip()

        if amount <= Decimal("0.00"):
            raise ValidationError("Payment amount must be greater than zero.")

        # Resolve payment account if not provided
        if not payment_account:
            if payment_method in ["BANK", "CARD", "CHEQUE"]:
                payment_account = Account.objects.filter(code="1021").first() or Account.objects.filter(parent__code="1020").first() or Account.objects.filter(code="1020").first()
            else:
                payment_account = Account.objects.filter(code="1011").first() or Account.objects.filter(parent__code="1010").first() or Account.objects.filter(code="1010").first()

        if payment_account.account_type != AccountType.ASSET:
            raise ValidationError(f"Payment account '{payment_account.name}' must be an Asset (Cash/Bank) account.")

        # Overpayment check
        balance_info = cls.get_customer_outstanding(customer.id)
        current_outstanding = balance_info["outstanding_balance"]

        if amount > current_outstanding:
            raise ValidationError(
                f"Payment amount (Rs. {amount:,.2f}) exceeds customer's outstanding receivable (Rs. {current_outstanding:,.2f}). Overpayment is not allowed."
            )

        payment_number = CustomerPayment.generate_payment_number(date)

        payment = CustomerPayment.objects.create(
            payment_number=payment_number,
            customer=customer,
            date=date,
            amount=amount,
            payment_method=payment_method,
            payment_account=payment_account,
            cheque_number=data.get("cheque_number", ""),
            cheque_date=data.get("cheque_date"),
            cheque_bank=data.get("cheque_bank", ""),
            reference=reference,
            notes=notes,
            status=CustomerPaymentStatus.DRAFT,
            created_by=user,
        )

        if submit_now:
            cls.submit_payment(payment, user)

        return payment

    @classmethod
    @transaction.atomic
    def submit_payment(cls, payment: CustomerPayment, user) -> CustomerPayment:
        """
        Submits a draft payment voucher and creates the balanced accounting entry:
        Debit: Payment Account (1010 Cash / 1020 Bank)
        Credit: Accounts Receivable (1030)
        """
        if payment.status != CustomerPaymentStatus.DRAFT:
            raise ValidationError(f"Cannot submit payment in '{payment.status}' status. Only DRAFT can be submitted.")

        # Accounts Receivable account
        ar_account = Account.objects.filter(code="1030").first() or Account.objects.get(code="1030")

        entry_lines = [
            {
                "account": payment.payment_account,
                "debit": payment.amount,
                "credit": Decimal("0.00"),
                "description": f"Payment received from {payment.customer.name} ({payment.payment_number})",
            },
            {
                "account": ar_account,
                "debit": Decimal("0.00"),
                "credit": payment.amount,
                "description": f"Receivable settlement for {payment.customer.name}",
            },
        ]

        journal_entry = AccountingService.create_journal_entry(
            entry_date=payment.date,
            reference_type=ReferenceType.CUSTOMER_PAYMENT,
            reference_id=payment.payment_number,
            narration=f"Customer Payment [{payment.payment_number}] - {payment.customer.name} (Amount: Rs. {payment.amount:,.2f})",
            lines=entry_lines,
            created_by=user,
            post_immediately=True,
        )

        payment.journal_entry = journal_entry
        payment.status = CustomerPaymentStatus.SUBMITTED
        payment.submitted_by = user
        payment.submitted_at = timezone.now()
        payment.save(update_fields=["journal_entry", "status", "submitted_by", "submitted_at", "updated_at"])

        # Reallocate customer payments across invoices in FIFO order
        cls.reallocate_customer_payments(payment.customer)

        return payment

    @classmethod
    @transaction.atomic
    def cancel_payment(cls, payment: CustomerPayment, user, reason: str = "") -> CustomerPayment:
        """
        Cancels a customer payment and generates a counter-reversal journal entry:
        Debit: Accounts Receivable (1030)
        Credit: Payment Account (1010 Cash / 1020 Bank)
        """
        if payment.status == CustomerPaymentStatus.CANCELLED:
            raise ValidationError("Payment is already cancelled.")

        if payment.status == CustomerPaymentStatus.SUBMITTED and payment.journal_entry:
            ar_account = Account.objects.filter(code="1030").first() or Account.objects.get(code="1030")

            reversal_lines = [
                {
                    "account": ar_account,
                    "debit": payment.amount,
                    "credit": Decimal("0.00"),
                    "description": f"Reversal of payment {payment.payment_number} - {reason or 'Payment Cancelled'}",
                },
                {
                    "account": payment.payment_account,
                    "debit": Decimal("0.00"),
                    "credit": payment.amount,
                    "description": f"Reversal of payment {payment.payment_number}",
                },
            ]

            reversal_entry = AccountingService.create_journal_entry(
                entry_date=timezone.now().date(),
                reference_type=ReferenceType.REVERSAL,
                reference_id=f"REV-{payment.payment_number}",
                narration=f"Cancellation Reversal for Customer Payment [{payment.payment_number}]: {reason or 'N/A'}",
                lines=reversal_lines,
                created_by=user,
                post_immediately=True,
            )
            payment.reversal_journal_entry = reversal_entry

        payment.status = CustomerPaymentStatus.CANCELLED
        payment.cancelled_by = user
        payment.cancelled_at = timezone.now()
        payment.cancellation_reason = reason
        payment.save(update_fields=["status", "reversal_journal_entry", "cancelled_by", "cancelled_at", "cancellation_reason", "updated_at"])

        # Reallocate customer payments across invoices
        cls.reallocate_customer_payments(payment.customer)

        return payment

    @classmethod
    def get_customer_statement(cls, customer_id: int, start_date=None, end_date=None) -> dict:
        """
        Generates comprehensive chronological statement of account for a customer:
        Opening Balance + Credit Sales (Debit) - Payments (Credit) - Sales Returns (Credit) = Closing Balance.
        """
        customer = Customer.objects.get(pk=customer_id)

        # Collect all transactional items
        events = []

        # 1. Sales Invoices and Counter Payments
        sales_qs = Sale.objects.filter(customer=customer, status=SaleStatus.COMPLETED)
        if start_date:
            sales_qs = sales_qs.filter(date__gte=start_date)
        if end_date:
            sales_qs = sales_qs.filter(date__lte=end_date)

        for sale in sales_qs:
            # A) Full Invoice Debit
            events.append({
                "date": sale.date,
                "created_at": sale.created_at,
                "type": "SALE",
                "type_display": "Sale Invoice",
                "reference": sale.invoice_number,
                "description": f"POS Sale Invoice ({sale.items.count()} items - Total: Rs. {sale.grand_total:,.2f})",
                "debit": float(sale.grand_total),
                "credit": 0.0,
            })

            # B) Itemized Counter / Upfront Payments (Cash, Card, Cheque, Split)
            if sale.payments.exists():
                for p in sale.payments.all():
                    if p.payment_method != PaymentMethodType.CREDIT and p.amount > Decimal("0.00"):
                        if p.payment_method == PaymentMethodType.CHEQUE:
                            method_label = "Cheque"
                            extra_desc = f" (Cheque #{sale.cheque_number})" if sale.cheque_number else ""
                        elif p.payment_method == PaymentMethodType.CARD:
                            method_label = "Card / Bank"
                            extra_desc = ""
                        else:
                            method_label = "Cash"
                            extra_desc = ""

                        events.append({
                            "date": sale.date,
                            "created_at": sale.created_at,
                            "type": "PAYMENT",
                            "type_display": f"Counter Payment ({method_label})",
                            "reference": sale.invoice_number,
                            "description": f"Immediate checkout settlement via {method_label}{extra_desc} for {sale.invoice_number}",
                            "debit": 0.0,
                            "credit": float(p.amount),
                        })
            else:
                upfront_paid = Decimal("0.00")
                method_label = ""
                extra_desc = ""
                if sale.payment_method == PaymentMethodType.CASH:
                    upfront_paid = sale.grand_total
                    method_label = "Cash"
                elif sale.payment_method == PaymentMethodType.CARD:
                    upfront_paid = sale.grand_total
                    method_label = "Card / Bank"
                elif sale.payment_method == PaymentMethodType.CHEQUE:
                    upfront_paid = sale.grand_total
                    method_label = "Cheque"
                    extra_desc = f" (Cheque #{sale.cheque_number})" if sale.cheque_number else ""
                elif sale.payment_method == PaymentMethodType.SPLIT:
                    upfront_paid = min(sale.paid_amount, sale.grand_total)
                    method_label = "Upfront Settlement"

                if upfront_paid > Decimal("0.00"):
                    events.append({
                        "date": sale.date,
                        "created_at": sale.created_at,
                        "type": "PAYMENT",
                        "type_display": f"Counter Payment ({method_label})",
                        "reference": sale.invoice_number,
                        "description": f"Immediate checkout settlement via {method_label}{extra_desc} for {sale.invoice_number}",
                        "debit": 0.0,
                        "credit": float(upfront_paid),
                    })

        # 2. Customer Payments (Payment Vouchers)
        payments_qs = CustomerPayment.objects.filter(customer=customer, status=CustomerPaymentStatus.SUBMITTED)
        if start_date:
            payments_qs = payments_qs.filter(date__gte=start_date)
        if end_date:
            payments_qs = payments_qs.filter(date__lte=end_date)

        for pay in payments_qs:
            account_name = pay.payment_account.name if pay.payment_account else "Payment Account"
            if pay.payment_method == "CHEQUE":
                type_display = f"Payment Voucher (Cheque #{pay.cheque_number})" if pay.cheque_number else "Payment Voucher (Cheque)"
                desc = pay.notes or f"Cheque #{pay.cheque_number} ({pay.cheque_bank or account_name}) settlement" if pay.cheque_number else f"Cheque settlement via {account_name}"
            elif pay.payment_method in ["BANK", "CARD"]:
                type_display = "Payment Voucher (Bank / Card)"
                desc = pay.notes or f"Bank / Card settlement via {account_name}"
            else:
                type_display = "Payment Voucher (Cash)"
                desc = pay.notes or f"Cash settlement via {account_name}"

            events.append({
                "date": pay.date,
                "created_at": pay.created_at,
                "type": "PAYMENT",
                "type_display": type_display,
                "reference": pay.payment_number,
                "description": desc,
                "debit": 0.0,
                "credit": float(pay.amount),
            })

        # 3. Sales Returns
        returns_qs = SalesReturn.objects.filter(original_sale__customer=customer)
        if start_date:
            returns_qs = returns_qs.filter(date__gte=start_date)
        if end_date:
            returns_qs = returns_qs.filter(date__lte=end_date)

        for ret in returns_qs:
            account_name = ret.payment_account.name if ret.payment_account else "Payment Account"
            is_ar_deduction = (ret.original_sale.payment_method == PaymentMethodType.CREDIT)

            if ret.refund_method == "CHEQUE":
                type_display = f"Sales Return (Cheque #{ret.cheque_number})" if ret.cheque_number else "Sales Return (Cheque)"
                desc = f"Refund paid via Cheque #{ret.cheque_number} ({ret.cheque_bank or account_name}) - {ret.reason or 'Sales Return'}" if ret.cheque_number else f"Refund paid via Cheque ({account_name}) - {ret.reason or 'Sales Return'}"
            elif ret.refund_method in ["BANK", "CARD"]:
                type_display = "Sales Return (Bank / Card)"
                desc = f"Refund paid via Bank Transfer ({account_name}) - {ret.reason or 'Sales Return'}"
            elif ret.refund_method == PaymentMethodType.CREDIT or is_ar_deduction:
                type_display = "Sales Return (Credit Note / AR Deduction)"
                desc = f"Return credited to receivable ({ret.original_sale.invoice_number}) - {ret.reason or 'Sales Return'}"
            else:
                type_display = "Sales Return (Cash)"
                desc = f"Refund paid in Cash ({account_name}) - {ret.reason or 'Sales Return'}"

            events.append({
                "date": ret.date,
                "created_at": ret.created_at,
                "type": "RETURN",
                "type_display": type_display,
                "reference": ret.return_number,
                "description": desc,
                "debit": 0.0,
                "credit": float(ret.refund_amount) if is_ar_deduction else 0.0,
                "is_direct_refund": not is_ar_deduction,
                "refund_amount": float(ret.refund_amount),
            })

        # Sort chronologically
        events.sort(key=lambda x: (x["date"], x["created_at"]))

        base_opening = customer.opening_balance or Decimal("0.00")
        running_balance = base_opening
        ledger_rows = []
        total_sales = Decimal("0.00")
        total_payments = Decimal("0.00")
        total_returns = Decimal("0.00")

        for ev in events:
            dr = Decimal(str(ev["debit"]))
            cr = Decimal(str(ev["credit"]))
            if ev["type"] == "SALE":
                total_sales += dr
                running_balance += dr
            elif ev["type"] == "PAYMENT":
                total_payments += cr
                running_balance -= cr
            elif ev["type"] == "RETURN":
                if not ev.get("is_direct_refund"):
                    total_returns += cr
                    running_balance -= cr

            ledger_rows.append({
                "date": str(ev["date"]),
                "type": ev["type"],
                "type_display": ev["type_display"],
                "reference": ev["reference"],
                "description": ev["description"],
                "debit": float(dr) if ev["type"] == "SALE" else 0.0,
                "credit": float(cr) if ev["type"] in ["PAYMENT", "RETURN"] else 0.0,
                "returned_amount": float(ev.get("refund_amount", cr)) if ev["type"] == "RETURN" else 0.0,
                "running_balance": float(max(Decimal("0.00"), running_balance)),
            })

        closing_balance = max(Decimal("0.00"), running_balance)

        ledger_rows.reverse()

        return {
            "customer": {
                "id": customer.id,
                "customer_id": customer.customer_id,
                "name": customer.name,
                "phone": customer.phone,
                "email": customer.email,
                "address": customer.address,
                "credit_enabled": customer.credit_enabled,
                "is_walkin": customer.is_walkin,
            },
            "period": {
                "start_date": str(start_date) if start_date else None,
                "end_date": str(end_date) if end_date else None,
            },
            "summary": {
                "opening_balance": float(base_opening),
                "total_debit": float(base_opening + total_sales),
                "total_sales": float(total_sales),
                "total_payments": float(total_payments),
                "total_returns": float(total_returns),
                "total_credit": float(total_payments + total_returns),
                "closing_balance": float(closing_balance),
            },
            "rows": ledger_rows,
        }

    @classmethod
    def get_receivables_report(cls, start_date=None, end_date=None, customer_id=None, status=None) -> dict:
        """
        Master analytical report for Customer Accounts Receivable and Outstanding aging.
        """
        customers_qs = Customer.objects.filter(is_walkin=False, is_active=True)
        if customer_id:
            customers_qs = customers_qs.filter(id=customer_id)

        rows = []
        grand_total_sales = Decimal("0.00")
        grand_total_returns = Decimal("0.00")
        grand_total_payments = Decimal("0.00")
        grand_total_outstanding = Decimal("0.00")

        for cust in customers_qs:
            info = cls.get_customer_outstanding(cust.id)
            sales = info["total_credit_sales"]
            returns = info.get("total_returns", Decimal("0.00"))
            payments = info["total_payments"]
            outstanding = info["outstanding_balance"]

            grand_total_sales += sales
            grand_total_returns += returns
            grand_total_payments += payments
            grand_total_outstanding += outstanding

            rows.append({
                "customer_id": cust.id,
                "customer_code": cust.customer_id,
                "name": cust.name,
                "phone": cust.phone or "-",
                "credit_enabled": cust.credit_enabled,
                "total_credit_sales": float(sales),
                "total_returns": float(returns),
                "total_payments": float(payments),
                "outstanding_balance": float(outstanding),
                "status": "Paid" if outstanding == 0 else "Outstanding",
            })

        return {
            "summary": {
                "total_registered_customers": customers_qs.count(),
                "total_credit_sales": float(grand_total_sales),
                "total_sales_returns": float(grand_total_returns),
                "net_credit_invoiced": float(grand_total_sales - grand_total_returns),
                "total_payments_collected": float(grand_total_payments),
                "total_outstanding_receivables": float(grand_total_outstanding),
            },
            "rows": rows,
        }

    @classmethod
    def bulk_import_customers(cls, rows: List[Dict[str, Any]], created_by=None) -> Dict[str, Any]:
        """
        Processes normalized customer rows from Excel or CSV.
        Auto-generates sequential customer_id if not provided and creates opening balance accounting entries.
        """
        total = len(rows)
        created_count = 0
        skipped_count = 0
        errors = []
        created_customers = []

        for index, row in enumerate(rows, start=1):
            name = str(row.get("name") or "").strip()
            if not name:
                errors.append(f"Row {index}: Missing required Customer Name.")
                skipped_count += 1
                continue

            customer_id = str(row.get("customer_id") or row.get("code") or "").strip().upper()
            if customer_id and Customer.objects.filter(customer_id__iexact=customer_id).exists():
                errors.append(f"Row {index}: Customer ID '{customer_id}' already exists. Skipped.")
                skipped_count += 1
                continue

            phone = str(row.get("phone") or "").strip()
            if phone and Customer.objects.filter(phone=phone).exists():
                errors.append(f"Row {index}: Customer with phone '{phone}' already exists. Skipped.")
                skipped_count += 1
                continue

            email = str(row.get("email") or "").strip() or None
            address = str(row.get("address") or "").strip() or None
            notes = str(row.get("notes") or row.get("description") or "").strip() or None

            credit_val = str(row.get("credit_enabled", "true")).strip().lower()
            credit_enabled = credit_val in ["true", "1", "yes", "y", "enabled"]

            try:
                opening_balance = Decimal(str(row.get("opening_balance") or "0.00"))
            except Exception:
                opening_balance = Decimal("0.00")

            try:
                with transaction.atomic():
                    customer = Customer(
                        customer_id=customer_id or Customer.generate_customer_id(),
                        name=name,
                        phone=phone or None,
                        email=email,
                        address=address,
                        credit_enabled=credit_enabled,
                        opening_balance=opening_balance,
                        notes=notes,
                        is_walkin=False,
                        is_active=True,
                    )
                    customer.save()

                    if customer.opening_balance > 0:
                        AccountingService.record_customer_opening_balance(
                            customer=customer,
                            amount=customer.opening_balance,
                            created_by=created_by,
                        )

                created_count += 1
                created_customers.append({
                    "id": customer.id,
                    "customer_id": customer.customer_id,
                    "name": customer.name,
                    "phone": customer.phone or "",
                    "credit_enabled": customer.credit_enabled,
                    "opening_balance": float(customer.opening_balance),
                })
            except Exception as e:
                errors.append(f"Row {index} ('{name}'): {str(e)}")
                skipped_count += 1

        return {
            "total_rows": total,
            "created_count": created_count,
            "skipped_count": skipped_count,
            "errors": errors,
            "created_customers": created_customers,
        }

    @classmethod
    def generate_customer_excel_template(cls) -> bytes:
        """
        Builds a styled sample Excel file matching Customer form fields.
        Customer ID is auto-generated upon import, exactly matching manual customer creation.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Import Template"

        headers = [
            "Customer Name *",
            "Phone Number",
            "Email Address",
            "Billing Address",
            "Credit Allowed (Yes/No)",
            "Opening Balance (Rs.)",
            "Notes / Remarks",
        ]
        ws.append(headers)

        sample_rows = [
            ["Ahmed Enterprises", "+92 300 1234567", "ahmed@example.com", "Shop 4, Commercial Plaza, Lahore", "Yes", 15000.00, "Regular wholesale buyer"],
            ["Fatima Super Store", "+92 321 7654321", "fatima@store.pk", "Main Market, Gulberg, Lahore", "Yes", 0.00, "Cash & Credit customer"],
            ["Usman Trader", "+92 333 9876543", "usman.traders@gmail.com", "DHA Phase 5, Lahore", "No", 0.00, "Walk-in cash only customer"],
        ]

        for r in sample_rows:
            ws.append(r)

        # Style headers
        header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = openpyxl.styles.Font(name="Arial", size=11, bold=True, color="FFFFFF")
        border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin", color="CCCCCC"),
            right=openpyxl.styles.Side(style="thin", color="CCCCCC"),
            top=openpyxl.styles.Side(style="thin", color="CCCCCC"),
            bottom=openpyxl.styles.Side(style="thin", color="CCCCCC"),
        )

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        for row_idx in range(2, len(sample_rows) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx == 6:  # Opening balance
                    cell.number_format = "#,##0.00"

        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()

    @classmethod
    def bulk_import_suppliers(cls, rows: List[Dict[str, Any]], created_by=None) -> Dict[str, Any]:
        """
        Processes normalized supplier rows from Excel or CSV.
        Auto-generates sequential supplier_id and creates opening balance accounting entries.
        """
        total = len(rows)
        created_count = 0
        skipped_count = 0
        errors = []
        created_suppliers = []

        for index, row in enumerate(rows, start=1):
            name = str(row.get("name") or "").strip()
            company_name = str(row.get("company_name") or row.get("company") or "").strip()

            if not name and not company_name:
                errors.append(f"Row {index}: Missing Supplier / Company Name.")
                skipped_count += 1
                continue

            if not name:
                name = company_name

            supplier_id = str(row.get("supplier_id") or row.get("code") or "").strip().upper()
            if supplier_id and Supplier.objects.filter(supplier_id__iexact=supplier_id).exists():
                errors.append(f"Row {index}: Supplier ID '{supplier_id}' already exists. Skipped.")
                skipped_count += 1
                continue

            phone = str(row.get("phone") or "").strip()
            if phone and Supplier.objects.filter(phone=phone).exists():
                errors.append(f"Row {index}: Supplier with phone '{phone}' already exists. Skipped.")
                skipped_count += 1
                continue

            email = str(row.get("email") or "").strip() or None
            address = str(row.get("address") or "").strip() or None
            tax_id = str(row.get("tax_id") or row.get("ntn") or row.get("strn") or "").strip() or None
            notes = str(row.get("notes") or row.get("description") or "").strip() or None

            try:
                opening_balance = Decimal(str(row.get("opening_balance") or "0.00"))
            except Exception:
                opening_balance = Decimal("0.00")

            try:
                with transaction.atomic():
                    supplier = Supplier(
                        supplier_id=supplier_id or Supplier.generate_supplier_id(),
                        name=name,
                        company_name=company_name or None,
                        phone=phone or None,
                        email=email,
                        address=address,
                        tax_id=tax_id,
                        opening_balance=opening_balance,
                        notes=notes,
                        is_active=True,
                    )
                    supplier.save()

                    if supplier.opening_balance > 0:
                        AccountingService.record_supplier_opening_balance(
                            supplier=supplier,
                            amount=supplier.opening_balance,
                            created_by=created_by,
                        )

                created_count += 1
                created_suppliers.append({
                    "id": supplier.id,
                    "supplier_id": supplier.supplier_id,
                    "name": supplier.name,
                    "company_name": supplier.company_name or "",
                    "phone": supplier.phone or "",
                    "tax_id": supplier.tax_id or "",
                    "opening_balance": float(supplier.opening_balance),
                })
            except Exception as e:
                errors.append(f"Row {index} ('{name}'): {str(e)}")
                skipped_count += 1

        return {
            "total_rows": total,
            "created_count": created_count,
            "skipped_count": skipped_count,
            "errors": errors,
            "created_suppliers": created_suppliers,
        }

    @classmethod
    def generate_supplier_excel_template(cls) -> bytes:
        """
        Builds a styled sample Excel file matching Supplier form fields.
        Supplier ID is auto-generated upon import, exactly matching manual supplier creation.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Supplier Import Template"

        headers = [
            "Contact Person Name *",
            "Company / Business Name",
            "Phone Number",
            "Email Address",
            "Office / Factory Address",
            "Tax / NTN / STRN",
            "Opening Payable Balance (Rs.)",
            "Notes / Payment Terms",
        ]
        ws.append(headers)

        sample_rows = [
            ["Kamran Sheikh", "Nestle Pakistan Ltd", "+92 42 111 637853", "orders@nestle.com.pk", "308 Upper Mall, Lahore", "NTN-0891234-1", 45000.00, "Net 15 days credit"],
            ["Saima Tariq", "Unilever Pakistan", "+92 21 111 864538", "supply@unilever.com", "Avari Plaza, Fatima Jinnah Road, Karachi", "NTN-0765432-8", 25000.00, "Direct distributor"],
            ["Zubair Hashmi", "National Foods Limited", "+92 21 35077001", "corporate@nfoods.com", "F-133, SITE, Karachi", "NTN-1423876-5", 0.00, "Spices & food condiments"],
        ]

        for r in sample_rows:
            ws.append(r)

        # Style headers
        header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = openpyxl.styles.Font(name="Arial", size=11, bold=True, color="FFFFFF")
        border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin", color="CCCCCC"),
            right=openpyxl.styles.Side(style="thin", color="CCCCCC"),
            top=openpyxl.styles.Side(style="thin", color="CCCCCC"),
            bottom=openpyxl.styles.Side(style="thin", color="CCCCCC"),
        )

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        for row_idx in range(2, len(sample_rows) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = border
                if col_idx == 8:  # Opening balance
                    cell.number_format = "#,##0.00"

        # Auto-adjust column width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

        stream = io.BytesIO()
        wb.save(stream)
        stream.seek(0)
        return stream.getvalue()
