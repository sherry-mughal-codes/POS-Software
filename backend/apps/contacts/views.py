import io
import csv
import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import models
from django.core.exceptions import ValidationError as DjangoValidationError
from rest_framework.exceptions import ValidationError as DRFValidationError

from apps.contacts.models import Customer, Supplier, CustomerPayment, CustomerPaymentStatus
from apps.contacts.serializers import (
    CustomerSerializer,
    SupplierSerializer,
    CustomerPaymentSerializer,
    CustomerPaymentCreateSerializer,
)
from apps.contacts.services import CustomerReceivableService
from apps.core.permissions import IsAdminOrManager


class CustomerViewSet(viewsets.ModelViewSet):
    """
    Customer Master Data API with phone search, credit eligibility, and auto-ID generator.
    """
    queryset = Customer.objects.all()
    serializer_class = CustomerSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ["destroy", "toggle_status"]:
            return [IsAdminOrManager()]
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get("search")
        credit_enabled = self.request.query_params.get("credit_enabled")
        is_active = self.request.query_params.get("is_active")

        if search:
            search = search.strip()
            qs = qs.filter(
                models.Q(name__icontains=search)
                | models.Q(phone__icontains=search)
                | models.Q(customer_id__icontains=search)
            )
        if credit_enabled is not None:
            qs = qs.filter(credit_enabled=credit_enabled.lower() == "true")
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")

        return qs

    def perform_create(self, serializer):
        customer = serializer.save()
        if customer.opening_balance > 0 and not customer.is_walkin:
            from apps.accounting.services import AccountingService
            AccountingService.record_customer_opening_balance(
                customer=customer,
                amount=customer.opening_balance,
                created_by=self.request.user if self.request.user.is_authenticated else None,
            )

    def perform_update(self, serializer):
        old_opening = serializer.instance.opening_balance
        customer = serializer.save()
        if not customer.is_walkin and customer.opening_balance != old_opening:
            from apps.accounting.services import AccountingService
            AccountingService.record_customer_opening_balance(
                customer=customer,
                amount=customer.opening_balance,
                created_by=self.request.user if self.request.user.is_authenticated else None,
            )

    def destroy(self, request, *args, **kwargs):
        customer = self.get_object()
        if customer.is_walkin:
            return Response(
                {"detail": "The default system Walk-in Customer cannot be deleted."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        self.perform_destroy(customer)
        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get"], url_path="next-id")
    def next_id(self, request):
        """Generates the next sequential customer identifier (e.g. CUS-0001, CUS-0002)."""
        return Response({"next_id": Customer.generate_customer_id()})

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        """
        Imports bulk customers from uploaded Excel/CSV file or JSON rows payload.
        """
        rows = []
        file_obj = request.FILES.get("file")

        if file_obj:
            filename = file_obj.name.lower()
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                try:
                    wb = openpyxl.load_workbook(file_obj, data_only=True)
                    ws = wb.active
                    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

                    field_map = {}
                    for idx, h in enumerate(headers):
                        if "code" in h or "id" in h:
                            field_map["customer_id"] = idx
                        elif "name" in h or "client" in h or "customer" in h:
                            field_map["name"] = idx
                        elif "phone" in h or "mobile" in h or "contact" in h or "cell" in h:
                            field_map["phone"] = idx
                        elif "email" in h or "mail" in h:
                            field_map["email"] = idx
                        elif "address" in h or "city" in h or "location" in h:
                            field_map["address"] = idx
                        elif "credit" in h or "allow" in h or "eligible" in h:
                            field_map["credit_enabled"] = idx
                        elif "open" in h or "balance" in h or "receivable" in h:
                            field_map["opening_balance"] = idx
                        elif "note" in h or "remark" in h or "desc" in h:
                            field_map["notes"] = idx

                    for row_cells in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row_cells):
                            continue
                        row_dict = {}
                        for field_name, col_idx in field_map.items():
                            val = row_cells[col_idx] if col_idx < len(row_cells) else None
                            row_dict[field_name] = str(val).strip() if val is not None else ""
                        if row_dict.get("name"):
                            rows.append(row_dict)
                except Exception as e:
                    return Response({"detail": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            elif filename.endswith(".csv"):
                try:
                    content = file_obj.read().decode("utf-8-sig")
                    reader = csv.DictReader(io.StringIO(content))
                    for r in reader:
                        cleaned = {k.strip().lower(): v for k, v in r.items() if k}
                        rows.append({
                            "name": cleaned.get("customer name") or cleaned.get("name") or cleaned.get("client name") or "",
                            "customer_id": cleaned.get("customer code") or cleaned.get("customer_id") or cleaned.get("code") or "",
                            "phone": cleaned.get("phone number") or cleaned.get("phone") or cleaned.get("mobile") or "",
                            "email": cleaned.get("email address") or cleaned.get("email") or "",
                            "address": cleaned.get("billing address") or cleaned.get("address") or "",
                            "credit_enabled": cleaned.get("credit allowed") or cleaned.get("credit_enabled") or "yes",
                            "opening_balance": cleaned.get("opening balance") or cleaned.get("opening_balance") or 0,
                            "notes": cleaned.get("notes") or cleaned.get("remarks") or "",
                        })
                except Exception as e:
                    return Response({"detail": f"Failed to parse CSV file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"detail": "Unsupported file format. Please upload .xlsx, .xls, or .csv"}, status=status.HTTP_400_BAD_REQUEST)
        elif isinstance(request.data, list):
            rows = request.data
        elif isinstance(request.data.get("items"), list):
            rows = request.data.get("items")
        elif isinstance(request.data.get("customers"), list):
            rows = request.data.get("customers")

        if not rows:
            return Response({"detail": "No valid customer rows found to import."}, status=status.HTTP_400_BAD_REQUEST)

        result = CustomerReceivableService.bulk_import_customers(rows, created_by=request.user)
        return Response(result, status=status.HTTP_200_OK if result["created_count"] > 0 else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """
        Downloads styled sample Excel template for bulk customer imports.
        """
        excel_bytes = CustomerReceivableService.generate_customer_excel_template()
        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Customer_Bulk_Import_Template.xlsx"'
        return response

    @action(detail=False, methods=["get"], url_path="walkin")
    def walkin(self, request):
        """Retrieves the default Walk-in Customer record for anonymous POS checkouts."""
        walkin = Customer.objects.filter(is_walkin=True).first()
        if not walkin:
            walkin = Customer.objects.create(
                customer_id="CUS-000001",
                name="Walk-in Customer",
                is_walkin=True,
                credit_enabled=False,
                is_active=True,
                notes="Default system record for anonymous counter sales",
            )
        return Response(CustomerSerializer(walkin).data)

    @action(detail=True, methods=["post"], url_path="toggle-status")
    def toggle_status(self, request, pk=None):
        """Soft-deactivates or reactivates a registered customer."""
        customer = self.get_object()
        if customer.is_walkin:
            return Response(
                {"detail": "The Walk-in Customer must always remain active."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        customer.is_active = not customer.is_active
        customer.save(update_fields=["is_active", "updated_at"])
        return Response({
            "id": customer.id,
            "name": customer.name,
            "is_active": customer.is_active,
            "detail": f"Customer '{customer.name}' is now {'active' if customer.is_active else 'inactive'}.",
        })

    @action(detail=True, methods=["get"], url_path="statement")
    def statement(self, request, pk=None):
        """Generates comprehensive chronological statement of account for customer."""
        customer = self.get_object()
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        statement_data = CustomerReceivableService.get_customer_statement(
            customer_id=customer.id,
            start_date=start_date,
            end_date=end_date,
        )
        return Response(statement_data, status=status.HTTP_200_OK)

    @action(detail=True, methods=["get"], url_path="outstanding")
    def outstanding(self, request, pk=None):
        """Returns authoritative current outstanding receivable balance calculation."""
        customer = self.get_object()
        data = CustomerReceivableService.get_customer_outstanding(customer.id)
        return Response(data, status=status.HTTP_200_OK)


class SupplierViewSet(viewsets.ModelViewSet):
    """
    Supplier Master Data API for purchasing and payables.
    """
    queryset = Supplier.objects.all()
    serializer_class = SupplierSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy", "toggle_status"]:
            return [IsAdminOrManager()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        supplier = serializer.save()
        if supplier.opening_balance > 0:
            from apps.accounting.services import AccountingService
            AccountingService.record_supplier_opening_balance(
                supplier=supplier,
                amount=supplier.opening_balance,
                created_by=self.request.user if self.request.user.is_authenticated else None,
            )

    def perform_update(self, serializer):
        old_opening = serializer.instance.opening_balance
        supplier = serializer.save()
        if supplier.opening_balance != old_opening:
            from apps.accounting.services import AccountingService
            AccountingService.record_supplier_opening_balance(
                supplier=supplier,
                amount=supplier.opening_balance,
                created_by=self.request.user if self.request.user.is_authenticated else None,
            )

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get("search")
        is_active = self.request.query_params.get("is_active")

        if search:
            search = search.strip()
            qs = qs.filter(
                models.Q(name__icontains=search)
                | models.Q(company_name__icontains=search)
                | models.Q(phone__icontains=search)
                | models.Q(supplier_id__icontains=search)
            )
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")

        return qs

    @action(detail=False, methods=["get"], url_path="next-id")
    def next_id(self, request):
        """Generates the next sequential supplier identifier (e.g. SUP-000001)."""
        return Response({"next_id": Supplier.generate_supplier_id()})

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        """
        Imports bulk suppliers from uploaded Excel/CSV file or JSON rows payload.
        """
        rows = []
        file_obj = request.FILES.get("file")

        if file_obj:
            filename = file_obj.name.lower()
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                try:
                    wb = openpyxl.load_workbook(file_obj, data_only=True)
                    ws = wb.active
                    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]

                    field_map = {}
                    for idx, h in enumerate(headers):
                        if "code" in h or "id" in h:
                            field_map["supplier_id"] = idx
                        elif "company" in h or "business" in h or "firm" in h or "vendor" in h:
                            field_map["company_name"] = idx
                        elif "contact" in h or "person" in h or "name" in h:
                            field_map["name"] = idx
                        elif "phone" in h or "mobile" in h or "tel" in h or "cell" in h:
                            field_map["phone"] = idx
                        elif "email" in h or "mail" in h:
                            field_map["email"] = idx
                        elif "address" in h or "city" in h or "factory" in h or "office" in h:
                            field_map["address"] = idx
                        elif "tax" in h or "ntn" in h or "strn" in h:
                            field_map["tax_id"] = idx
                        elif "open" in h or "balance" in h or "payable" in h:
                            field_map["opening_balance"] = idx
                        elif "note" in h or "remark" in h or "term" in h or "desc" in h:
                            field_map["notes"] = idx

                    for row_cells in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row_cells):
                            continue
                        row_dict = {}
                        for field_name, col_idx in field_map.items():
                            val = row_cells[col_idx] if col_idx < len(row_cells) else None
                            row_dict[field_name] = str(val).strip() if val is not None else ""
                        if row_dict.get("name") or row_dict.get("company_name"):
                            rows.append(row_dict)
                except Exception as e:
                    return Response({"detail": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            elif filename.endswith(".csv"):
                try:
                    content = file_obj.read().decode("utf-8-sig")
                    reader = csv.DictReader(io.StringIO(content))
                    for r in reader:
                        cleaned = {k.strip().lower(): v for k, v in r.items() if k}
                        rows.append({
                            "company_name": cleaned.get("company / business name") or cleaned.get("company name") or cleaned.get("company") or "",
                            "name": cleaned.get("contact person name") or cleaned.get("contact person") or cleaned.get("name") or "",
                            "supplier_id": cleaned.get("supplier code") or cleaned.get("supplier_id") or cleaned.get("code") or "",
                            "phone": cleaned.get("phone number") or cleaned.get("phone") or cleaned.get("mobile") or "",
                            "email": cleaned.get("email address") or cleaned.get("email") or "",
                            "address": cleaned.get("office / factory address") or cleaned.get("address") or "",
                            "tax_id": cleaned.get("tax / ntn / strn") or cleaned.get("tax_id") or cleaned.get("ntn") or "",
                            "opening_balance": cleaned.get("opening payable balance") or cleaned.get("opening_balance") or 0,
                            "notes": cleaned.get("notes / payment terms") or cleaned.get("notes") or "",
                        })
                except Exception as e:
                    return Response({"detail": f"Failed to parse CSV file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"detail": "Unsupported file format. Please upload .xlsx, .xls, or .csv"}, status=status.HTTP_400_BAD_REQUEST)
        elif isinstance(request.data, list):
            rows = request.data
        elif isinstance(request.data.get("items"), list):
            rows = request.data.get("items")
        elif isinstance(request.data.get("suppliers"), list):
            rows = request.data.get("suppliers")

        if not rows:
            return Response({"detail": "No valid supplier rows found to import."}, status=status.HTTP_400_BAD_REQUEST)

        result = CustomerReceivableService.bulk_import_suppliers(rows, created_by=request.user)
        return Response(result, status=status.HTTP_200_OK if result["created_count"] > 0 else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """
        Downloads styled sample Excel template for bulk supplier imports.
        """
        excel_bytes = CustomerReceivableService.generate_supplier_excel_template()
        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Supplier_Bulk_Import_Template.xlsx"'
        return response

    @action(detail=True, methods=["post"], url_path="toggle-status")
    def toggle_status(self, request, pk=None):
        """Soft-deactivates or reactivates a supplier."""
        supplier = self.get_object()
        supplier.is_active = not supplier.is_active
        supplier.save(update_fields=["is_active", "updated_at"])
        return Response({
            "id": supplier.id,
            "name": supplier.name,
            "is_active": supplier.is_active,
            "detail": f"Supplier '{supplier.name}' is now {'active' if supplier.is_active else 'inactive'}.",
        })


class CustomerPaymentViewSet(viewsets.ModelViewSet):
    """
    CRUD and workflow operations for Customer Payments against Accounts Receivable.
    """
    queryset = CustomerPayment.objects.all().select_related("customer", "payment_account", "created_by", "submitted_by", "cancelled_by")
    permission_classes = [IsAuthenticated]

    def get_serializer_class(self):
        if self.action == "create":
            return CustomerPaymentCreateSerializer
        return CustomerPaymentSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        customer_id = self.request.query_params.get("customer")
        status_filter = self.request.query_params.get("status")
        payment_method = self.request.query_params.get("payment_method")
        date_from = self.request.query_params.get("date_from")
        date_to = self.request.query_params.get("date_to")
        search = self.request.query_params.get("search")

        if customer_id:
            qs = qs.filter(customer_id=customer_id)
        if status_filter:
            qs = qs.filter(status=status_filter)
        if payment_method:
            qs = qs.filter(payment_method=payment_method)
        if date_from:
            qs = qs.filter(date__gte=date_from)
        if date_to:
            qs = qs.filter(date__lte=date_to)
        if search:
            qs = qs.filter(
                models.Q(payment_number__icontains=search)
                | models.Q(customer__name__icontains=search)
                | models.Q(reference__icontains=search)
            )

        return qs

    def create(self, request, *args, **kwargs):
        serializer = CustomerPaymentCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        submit_now = serializer.validated_data.pop("submit_now", True)

        try:
            payment = CustomerReceivableService.create_payment(
                data=serializer.validated_data,
                user=request.user,
                submit_now=submit_now,
            )
            output_serializer = CustomerPaymentSerializer(payment)
            return Response(output_serializer.data, status=status.HTTP_201_CREATED)
        except (DjangoValidationError, DRFValidationError) as e:
            return Response({"detail": str(e.message if hasattr(e, 'message') else e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"], url_path="submit")
    def submit(self, request, pk=None):
        """Submits a draft payment voucher and creates Accounts Receivable journal entry."""
        payment = self.get_object()
        try:
            submitted_payment = CustomerReceivableService.submit_payment(payment=payment, user=request.user)
            return Response(CustomerPaymentSerializer(submitted_payment).data, status=status.HTTP_200_OK)
        except (DjangoValidationError, DRFValidationError) as e:
            return Response({"detail": str(e.message if hasattr(e, 'message') else e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        """Cancels a payment voucher and posts Accounts Receivable counter-reversal."""
        payment = self.get_object()
        reason = request.data.get("reason", "").strip()
        try:
            cancelled_payment = CustomerReceivableService.cancel_payment(payment=payment, user=request.user, reason=reason)
            return Response(CustomerPaymentSerializer(cancelled_payment).data, status=status.HTTP_200_OK)
        except (DjangoValidationError, DRFValidationError) as e:
            return Response({"detail": str(e.message if hasattr(e, 'message') else e)}, status=status.HTTP_400_BAD_REQUEST)


class CustomerReceivablesReportView(APIView):
    """
    Master analytical report for Customer Accounts Receivable and Outstanding balances.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        start_date = request.query_params.get("start_date")
        end_date = request.query_params.get("end_date")
        customer_id = request.query_params.get("customer")
        status_param = request.query_params.get("status")

        report_data = CustomerReceivableService.get_receivables_report(
            start_date=start_date,
            end_date=end_date,
            customer_id=int(customer_id) if customer_id else None,
            status=status_param,
        )
        return Response(report_data, status=status.HTTP_200_OK)
