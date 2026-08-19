"""
Product Service Layer.
Handles SKU generation, atomic product creation with opening stock, and bulk Excel/CSV product imports.
"""

from decimal import Decimal
import io
import re
from typing import List, Dict, Any, Optional
from django.db import transaction, models
from django.core.exceptions import ValidationError
import openpyxl

from apps.products.models import Product, Category, Unit
from apps.inventory.models import StockMovement, MovementType


class ProductService:
    """
    Service layer for product master records, SKU generation, and bulk imports.
    """

    @classmethod
    def generate_sku(cls) -> str:
        """
        Generates the next sequential product SKU (e.g. PRD-00001, PRD-00002).
        """
        prefix = "PRD-"
        last = Product.objects.filter(sku__startswith=prefix).order_by("-id").first()
        if last:
            try:
                seq = int(last.sku.split("-")[-1]) + 1
            except (ValueError, IndexError):
                seq = Product.objects.count() + 1
        else:
            seq = Product.objects.count() + 1
        return f"{prefix}{seq:05d}"

    @classmethod
    @transaction.atomic
    def create_product(
        cls,
        name: str,
        category: Category,
        unit: Unit,
        purchase_price: Decimal = Decimal("0.00"),
        selling_price: Decimal = Decimal("0.00"),
        sku: Optional[str] = None,
        barcode: Optional[str] = None,
        min_stock_level: Decimal = Decimal("10.00"),
        opening_stock: Decimal = Decimal("0.00"),
        image_url: Optional[str] = None,
        description: Optional[str] = None,
        is_active: bool = True,
        created_by=None,
    ) -> Product:
        """
        Creates a product and initializes Opening Stock in the inventory ledger if specified.
        """
        if not sku or not sku.strip():
            sku = cls.generate_sku()
        else:
            sku = sku.strip().upper()

        if barcode:
            barcode = barcode.strip()
            if not barcode:
                barcode = None

        product = Product.objects.create(
            sku=sku,
            name=name.strip(),
            barcode=barcode,
            category=category,
            unit=unit,
            purchase_price=Decimal(str(purchase_price or "0.00")),
            selling_price=Decimal(str(selling_price or "0.00")),
            min_stock_level=Decimal(str(min_stock_level or "10.00")),
            image_url=image_url or "",
            description=description or "",
            is_active=is_active,
        )

        # Record Opening Stock in inventory ledger if opening_stock > 0
        opn_qty = Decimal(str(opening_stock or "0.00"))
        if opn_qty > Decimal("0.00"):
            StockMovement.objects.create(
                product=product,
                movement_type=MovementType.OPENING_STOCK,
                quantity=opn_qty,
                unit_cost=product.purchase_price,
                balance_after=opn_qty,
                reference_type="OPENING_BALANCE",
                reference_id=f"OPN-{product.sku}",
                notes=f"Initial Opening Stock for {product.name}",
                created_by=created_by,
            )

            # Auto-post double-entry GL entry to Inventory Asset (1040) and Owner's Equity (3010)
            valuation = opn_qty * product.purchase_price
            if valuation > Decimal("0.00"):
                try:
                    from apps.accounting.models import Account, ReferenceType
                    from apps.accounting.services import AccountingService
                    from django.utils import timezone

                    inv_acc = Account.objects.get(code="1040")
                    equity_acc = Account.objects.get(code="3010")
                    AccountingService.create_journal_entry(
                        entry_date=timezone.now().date(),
                        reference_type=ReferenceType.OPENING_BALANCE,
                        reference_id=f"OPN-{product.sku}",
                        lines=[
                            {
                                "account": inv_acc,
                                "debit": valuation,
                                "credit": Decimal("0.00"),
                                "description": f"Opening inventory asset for {product.name} ({opn_qty} @ Rs. {product.purchase_price:.2f})",
                            },
                            {
                                "account": equity_acc,
                                "debit": Decimal("0.00"),
                                "credit": valuation,
                                "description": f"Opening capital equity from product inventory ({product.name})",
                            },
                        ],
                        narration=f"Opening Stock Setup: {product.name} ({opn_qty} units @ Rs. {product.purchase_price:.2f})",
                        created_by=created_by,
                    )
                except Exception as e:
                    import logging
                    logging.getLogger(__name__).warning(f"Could not auto-post opening stock GL entry for {product.sku}: {e}")

        return product

    @classmethod
    @transaction.atomic
    def bulk_import_products(cls, rows: List[Dict[str, Any]], created_by=None) -> Dict[str, Any]:
        """
        Processes normalized product rows from Excel or CSV.
        Auto-creates missing categories and units when encountered.
        """
        total = len(rows)
        created_count = 0
        skipped_count = 0
        errors = []
        created_products = []

        # Default fallback category and unit
        default_cat, _ = Category.objects.get_or_create(
            code="GEN",
            defaults={"name": "General Products", "description": "Default category"}
        )
        default_unit, _ = Unit.objects.get_or_create(
            short_code="pcs",
            defaults={"name": "Piece", "allow_decimal": False}
        )

        for index, row in enumerate(rows, start=1):
            name = str(row.get("name") or "").strip()
            if not name:
                errors.append(f"Row {index}: Missing required Product Name.")
                skipped_count += 1
                continue

            sku = str(row.get("sku") or "").strip().upper()
            if sku and Product.objects.filter(sku__iexact=sku).exists():
                errors.append(f"Row {index}: SKU '{sku}' already exists. Skipped.")
                skipped_count += 1
                continue

            barcode = str(row.get("barcode") or "").strip()
            if barcode and Product.objects.filter(barcode=barcode).exists():
                errors.append(f"Row {index}: Barcode '{barcode}' already exists for another product. Skipped.")
                skipped_count += 1
                continue
            if not barcode:
                barcode = None

            # Category resolution / auto-creation
            cat_name = str(row.get("category") or "").strip()
            if cat_name:
                cat_code = re.sub(r"[^A-Z0-9]", "", cat_name.upper())[:10] or "CAT"
                category, _ = Category.objects.get_or_create(
                    name__iexact=cat_name,
                    defaults={"code": cat_code, "name": cat_name}
                )
            else:
                category = default_cat

            # Unit resolution / auto-creation
            unit_val = str(row.get("unit") or "").strip().lower()
            if unit_val:
                unit = Unit.objects.filter(
                    models.Q(short_code__iexact=unit_val) | models.Q(name__iexact=unit_val)
                ).first()
                if not unit:
                    unit = Unit.objects.create(
                        name=unit_val.title(),
                        short_code=unit_val[:10],
                        allow_decimal=unit_val in ["kg", "liter", "ltr", "gm", "gram", "meter", "m"]
                    )
            else:
                unit = default_unit

            try:
                purchase_price = Decimal(str(row.get("purchase_price") or "0.00"))
            except Exception:
                purchase_price = Decimal("0.00")

            try:
                selling_price = Decimal(str(row.get("selling_price") or "0.00"))
            except Exception:
                selling_price = Decimal("0.00")

            try:
                min_stock = Decimal(str(row.get("min_stock_level") or "10.00"))
            except Exception:
                min_stock = Decimal("10.00")

            try:
                opening_stock = Decimal(str(row.get("opening_stock") or row.get("quantity") or "0.00"))
            except Exception:
                opening_stock = Decimal("0.00")

            try:
                product = cls.create_product(
                    name=name,
                    category=category,
                    unit=unit,
                    purchase_price=purchase_price,
                    selling_price=selling_price,
                    sku=sku or None,
                    barcode=barcode,
                    min_stock_level=min_stock,
                    opening_stock=opening_stock,
                    description=str(row.get("description") or ""),
                    created_by=created_by,
                )
                created_count += 1
                created_products.append({
                    "id": product.id,
                    "sku": product.sku,
                    "name": product.name,
                    "category": product.category.name,
                    "unit": product.unit.name,
                    "purchase_price": float(product.purchase_price),
                    "selling_price": float(product.selling_price),
                    "opening_stock": float(opening_stock),
                })
            except Exception as e:
                errors.append(f"Row {index} ('{name}'): {str(e)}")
                skipped_count += 1

        return {
            "total_rows": total,
            "created_count": created_count,
            "skipped_count": skipped_count,
            "errors": errors,
            "created_products": created_products,
        }

    @classmethod
    def generate_excel_template(cls) -> bytes:
        """
        Builds a styled sample Excel file with headers and example product rows for user download.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Product Import Template"

        headers = [
            "Product Name *",
            "SKU (Optional)",
            "Barcode (Optional)",
            "Category",
            "Unit (e.g. pcs, kg)",
            "Purchase Price (Rs.)",
            "Selling Price (Rs.)",
            "Opening Quantity",
            "Min Stock Level",
            "Description",
        ]
        ws.append(headers)

        sample_rows = [
            ["Nestle Everyday Milk Powder 1kg", "PRD-00101", "8964000123456", "Dairy & Milk", "pcs", 1450.00, 1600.00, 50, 10, "1kg pouch pack"],
            ["Olpers Full Cream Milk 1 Liter", "PRD-00102", "8964000123457", "Dairy & Milk", "pcs", 270.00, 295.00, 120, 24, "1000ml Tetra pack"],
            ["Dalda Cooking Oil 5 Liter", "PRD-00103", "8964000123458", "Edible Oil & Ghee", "bottle", 2650.00, 2850.00, 30, 5, "5L Can with handle"],
            ["Tapal Danedar Black Tea 450g", "PRD-00104", "8964000123459", "Beverages", "pcs", 620.00, 680.00, 45, 10, "450g Hard Pack"],
            ["National Himalayan Pink Salt 800g", "PRD-00105", "8964000123460", "Spices & Salt", "pcs", 95.00, 120.00, 80, 15, "800g jar"],
        ]

        for r in sample_rows:
            ws.append(r)

        # Style header
        header_fill = openpyxl.styles.PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = openpyxl.styles.Font(name="Calibri", size=11, bold=True, color="FFFFFF")

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 24

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()
