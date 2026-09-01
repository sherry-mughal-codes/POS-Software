"""
API views for Product Master, Categories, and Units of Measure.
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db import models

from apps.products.models import Category, Unit, Product
from apps.products.serializers import (
    CategorySerializer,
    UnitSerializer,
    ProductSerializer,
)
from apps.core.permissions import IsAdminOrManager


class CategoryViewSet(viewsets.ModelViewSet):
    """
    Product Category Management API.
    """
    queryset = Category.objects.all().select_related("parent").prefetch_related("products", "children")
    serializer_class = CategorySerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAdminOrManager()]
        return [IsAuthenticated()]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.products.exists():
            return Response(
                {"detail": f"Cannot delete category '{instance.name}' because it contains {instance.products.count()} associated products. Reassign or deactivate them first."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if instance.children.exists():
            return Response(
                {"detail": f"Cannot delete category '{instance.name}' because it has subcategories."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


class UnitViewSet(viewsets.ModelViewSet):
    """
    Units of Measurement Management API.
    """
    queryset = Unit.objects.all().prefetch_related("products")
    serializer_class = UnitSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy"]:
            return [IsAdminOrManager()]
        return [IsAuthenticated()]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        if instance.products.exists():
            return Response(
                {"detail": f"Cannot delete unit '{instance.name}' because it is assigned to {instance.products.count()} products."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        self.perform_destroy(instance)
        return Response(status=status.HTTP_204_NO_CONTENT)


import csv
import io
import openpyxl
from django.http import HttpResponse
from apps.products.services import ProductService

class ProductViewSet(viewsets.ModelViewSet):
    """
    Product Master Catalog API with search, barcode lookup, category filtering, and soft-deactivation.
    """
    queryset = Product.objects.all().select_related("category", "unit")
    serializer_class = ProductSerializer
    permission_classes = [IsAuthenticated]

    def get_permissions(self):
        if self.action in ["create", "update", "partial_update", "destroy", "toggle_status", "bulk_import"]:
            return [IsAdminOrManager()]
        return [IsAuthenticated()]

    def get_queryset(self):
        qs = super().get_queryset()
        search = self.request.query_params.get("search")
        category_id = self.request.query_params.get("category")
        unit_id = self.request.query_params.get("unit")
        is_active = self.request.query_params.get("is_active")

        if search:
            search = search.strip()
            qs = qs.filter(
                models.Q(name__icontains=search)
                | models.Q(sku__icontains=search)
                | models.Q(barcode__iexact=search)
            )
        if category_id:
            qs = qs.filter(category_id=category_id)
        if unit_id:
            qs = qs.filter(unit_id=unit_id)
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")

        return qs

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        reasons = []

        # 1. Check current inventory stock
        current_stock = instance.get_current_stock()
        if current_stock > 0:
            unit_code = instance.unit.short_code if instance.unit else "units"
            reasons.append(f"Active stock on hand ({current_stock} {unit_code})")

        # 2. Check sales history
        from apps.sales.models import SaleItem
        sale_count = SaleItem.objects.filter(product=instance).count()
        if sale_count > 0:
            reasons.append(f"{sale_count} sales transaction(s)")

        # 3. Check purchase history
        from apps.purchases.models import PurchaseItem
        purchase_count = PurchaseItem.objects.filter(product=instance).count()
        if purchase_count > 0:
            reasons.append(f"{purchase_count} purchase order(s)")

        # 4. Check stock movements
        from apps.inventory.models import StockMovement, StockAdjustmentItem
        movement_count = StockMovement.objects.filter(product=instance).count()
        if movement_count > 0:
            reasons.append(f"{movement_count} stock movement record(s)")

        # 5. Check stock adjustments
        adj_count = StockAdjustmentItem.objects.filter(product=instance).count()
        if adj_count > 0:
            reasons.append(f"{adj_count} stock adjustment(s)")

        # 6. Check warranty claims
        from apps.warranty.models import CustomerWarrantyClaim, SupplierWarrantyClaimItem
        warranty_count = CustomerWarrantyClaim.objects.filter(
            models.Q(claimed_product=instance) | models.Q(replacement_product=instance)
        ).count()
        if warranty_count > 0:
            reasons.append(f"{warranty_count} customer warranty claim(s)")

        supplier_claim_count = SupplierWarrantyClaimItem.objects.filter(product=instance).count()
        if supplier_claim_count > 0:
            reasons.append(f"{supplier_claim_count} supplier warranty claim(s)")

        if reasons:
            reasons_str = ", ".join(reasons)
            return Response(
                {
                    "detail": f"Cannot delete product '{instance.name}' ({instance.sku}) because it has existing inventory or transactional records: {reasons_str}. You can deactivate this product instead to hide it from sales while preserving audit integrity."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Free of all inventory and transactions - delete permanently
        self.perform_destroy(instance)
        return Response(
            {"detail": f"Product '{instance.name}' has been permanently deleted."},
            status=status.HTTP_204_NO_CONTENT,
        )

    @action(detail=False, methods=["get"], url_path="lookup-barcode")
    def lookup_barcode(self, request):
        """Instant exact barcode lookup for POS scanner input."""
        barcode = request.query_params.get("barcode", "").strip()
        if not barcode:
            return Response({"detail": "Barcode parameter is required."}, status=status.HTTP_400_BAD_REQUEST)

        product = Product.objects.filter(barcode=barcode).first()
        if not product:
            return Response({"detail": f"No product found for barcode '{barcode}'."}, status=status.HTTP_404_NOT_FOUND)

        return Response(ProductSerializer(product).data)

    @action(detail=False, methods=["get"], url_path="next-sku")
    def next_sku(self, request):
        """Generates the next sequential recommended SKU (e.g. PRD-00029)."""
        return Response({"next_sku": ProductService.generate_sku()})

    @action(detail=False, methods=["post"], url_path="bulk-import")
    def bulk_import(self, request):
        """
        Imports bulk products from uploaded Excel/CSV file or JSON row array.
        """
        rows = []
        file_obj = request.FILES.get("file")

        if file_obj:
            filename = file_obj.name.lower()
            if filename.endswith(".xlsx") or filename.endswith(".xls"):
                try:
                    wb = openpyxl.load_workbook(file_obj, data_only=True)
                    ws = wb.active
                    headers = [str(cell.value or "").strip().lower() for cell in ws[1]]
                    
                    # Map standard header variations
                    field_map = {}
                    for idx, h in enumerate(headers):
                        if "product" in h or "name" in h or "title" in h or "item" in h:
                            field_map["name"] = idx
                        elif "sku" in h or "code" in h:
                            field_map["sku"] = idx
                        elif "barcode" in h or "ean" in h or "upc" in h:
                            field_map["barcode"] = idx
                        elif "cat" in h or "department" in h or "group" in h:
                            field_map["category"] = idx
                        elif "unit" in h or "uom" in h:
                            field_map["unit"] = idx
                        elif "purchase" in h or "cost" in h or "buy" in h:
                            field_map["purchase_price"] = idx
                        elif "sell" in h or "retail" in h or "price" in h:
                            field_map["selling_price"] = idx
                        elif "open" in h or "qty" in h or "quantity" in h or "stock" in h:
                            field_map["opening_stock"] = idx
                        elif "min" in h or "alert" in h or "threshold" in h:
                            field_map["min_stock_level"] = idx
                        elif "desc" in h or "note" in h or "detail" in h:
                            field_map["description"] = idx

                    for row_cells in ws.iter_rows(min_row=2, values_only=True):
                        if not any(row_cells):
                            continue
                        row_dict = {}
                        for field_name, col_idx in field_map.items():
                            val = row_cells[col_idx] if col_idx < len(row_cells) else None
                            row_dict[field_name] = str(val).strip() if val is not None else ""
                        if row_dict.get("name"):
                            rows.append(row_dict)
                except Exception as e:
                    return Response({"detail": f"Failed to parse Excel file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            elif filename.endswith(".csv"):
                try:
                    content = file_obj.read().decode("utf-8-sig")
                    reader = csv.DictReader(io.StringIO(content))
                    for r in reader:
                        cleaned = {k.strip().lower(): v for k, v in r.items() if k}
                        rows.append({
                            "name": cleaned.get("product name") or cleaned.get("name") or cleaned.get("title") or "",
                            "sku": cleaned.get("sku") or cleaned.get("code") or "",
                            "barcode": cleaned.get("barcode") or cleaned.get("ean") or "",
                            "category": cleaned.get("category") or "",
                            "unit": cleaned.get("unit") or cleaned.get("uom") or "pcs",
                            "purchase_price": cleaned.get("purchase price") or cleaned.get("purchase_price") or cleaned.get("cost") or 0,
                            "selling_price": cleaned.get("selling price") or cleaned.get("selling_price") or cleaned.get("price") or 0,
                            "opening_stock": cleaned.get("opening quantity") or cleaned.get("opening_stock") or cleaned.get("quantity") or 0,
                            "min_stock_level": cleaned.get("min stock level") or cleaned.get("min_stock_level") or 10,
                            "description": cleaned.get("description") or "",
                        })
                except Exception as e:
                    return Response({"detail": f"Failed to parse CSV file: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
            else:
                return Response({"detail": "Unsupported file format. Please upload .xlsx, .xls, or .csv"}, status=status.HTTP_400_BAD_REQUEST)
        elif isinstance(request.data, list):
            rows = request.data
        elif isinstance(request.data.get("items"), list):
            rows = request.data.get("items")
        elif isinstance(request.data.get("products"), list):
            rows = request.data.get("products")

        if not rows:
            return Response({"detail": "No valid product rows found to import."}, status=status.HTTP_400_BAD_REQUEST)

        result = ProductService.bulk_import_products(rows, created_by=request.user)
        return Response(result, status=status.HTTP_200_OK if result["created_count"] > 0 else status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=["get"], url_path="import-template")
    def import_template(self, request):
        """
        Downloads styled sample Excel template for bulk product imports.
        """
        excel_bytes = ProductService.generate_excel_template()
        response = HttpResponse(
            excel_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="Product_Bulk_Import_Template.xlsx"'
        return response

    @action(detail=True, methods=["post"], url_path="toggle-status")
    def toggle_status(self, request, pk=None):
        """Soft-deactivates or reactivates a product without deleting historical records."""
        product = self.get_object()
        product.is_active = not product.is_active
        product.save(update_fields=["is_active", "updated_at"])

        return Response({
            "id": product.id,
            "name": product.name,
            "is_active": product.is_active,
            "detail": f"Product '{product.name}' is now {'active' if product.is_active else 'inactive'}.",
        })
